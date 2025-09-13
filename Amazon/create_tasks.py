import pandas as pd
from datetime import datetime

# Create comprehensive macro tasks for Amazon website
tasks = [
    # ===== GENERAL TASKS (15) =====
    {
        "Task_ID": "T001",
        "Task_Name": "Advanced Multi-Category Product Analysis",
        "Category": "General",
        "Difficulty": "Hard",
        "Description": "Navigate through all 8 categories (electronics, books, clothing, home, sports, automotive, bestsellers, new-releases), collect product data including prices, ratings, and seller information, then generate a comparative analysis report showing average price per category and top-rated products in each category.",
        "Target_Elements": ".nav-links a, .product-card, .product-title, .product-price, .product-rating, .seller-info",
        "Required_Actions": "1. Click each category link 2. Wait for products to load 3. Extract all product data 4. Apply sorting filters 5. Collect statistics 6. Generate comparison report",
        "Success_Criteria": "Successfully collect data from all 8 categories, identify highest/lowest priced items per category, calculate average ratings, and export structured data with at least 50+ products",
        "Estimated_Time": "3-4 minutes",
        "Notes": "Must handle dynamic loading states and ensure complete data collection before moving to next category"
    },
    {
        "Task_ID": "T002", 
        "Task_Name": "Dynamic Shopping Cart Stress Test",
        "Category": "General",
        "Difficulty": "Expert",
        "Description": "Add 15+ different products to cart with varying quantities (1-5 each), modify quantities multiple times, remove specific items, calculate running totals, test cart persistence across page navigation, and simulate checkout process while tracking all state changes.",
        "Target_Elements": ".add-to-cart, .cart, #cartModal, .quantity-btn, .remove-item, #cartCount, #cartTotal",
        "Required_Actions": "1. Add products with different quantities 2. Open/close cart modal repeatedly 3. Modify quantities systematically 4. Remove items in specific order 5. Navigate pages while maintaining cart state 6. Verify localStorage persistence",
        "Success_Criteria": "Cart maintains accurate count and total through all operations, localStorage correctly updates, cart survives page navigation, and final checkout preparation is successful",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Test edge cases like quantity 0, maximum items, and rapid successive operations"
    },
    {
        "Task_ID": "T003",
        "Task_Name": "Comprehensive Jun Profile Data Mining", 
        "Category": "General",
        "Difficulty": "Medium",
        "Description": "Extract complete information about Jun's account including profile stats, all 3 products he sells (with full details), all reviews he has written (with ratings and dates), order history analysis, and create a seller performance report.",
        "Target_Elements": ".profile-page, .profile-stats, .stat-number, #junProductsGrid, .jun-review, .reviews-list, .orders-list",
        "Required_Actions": "1. Navigate to Jun's profile 2. Extract profile statistics 3. Collect all Jun's product details 4. Find and extract all Jun's reviews 5. Analyze order patterns 6. Generate seller report",
        "Success_Criteria": "Complete data extraction of Jun's 3 products, all written reviews with dates and ratings, profile statistics (orders, reviews, seller rating), and structured seller performance analysis",
        "Estimated_Time": "2-3 minutes", 
        "Notes": "Jun's reviews appear on multiple product pages, requires navigation to find all instances"
    },
    {
        "Task_ID": "T004",
        "Task_Name": "Advanced Search Autocomplete Intelligence Gathering",
        "Category": "General", 
        "Difficulty": "Hard",
        "Description": "Systematically test search autocomplete by inputting partial terms for different product categories, collect all suggestion responses, analyze suggestion patterns, test edge cases with special characters, and create a comprehensive autocomplete behavior map.",
        "Target_Elements": "#searchInput, .search-suggestions, .suggestion-item, .search-category",
        "Required_Actions": "1. Input various partial search terms 2. Capture autocomplete suggestions 3. Test different categories 4. Try special characters and edge cases 5. Analyze suggestion algorithms 6. Map autocomplete behavior patterns",
        "Success_Criteria": "Collect autocomplete data for 20+ different search terms across all categories, identify suggestion patterns, test at least 5 edge cases, and document complete autocomplete behavior",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Must handle debounced search suggestions (300ms delay) and test both search history and product title suggestions"
    },
    {
        "Task_ID": "T005",
        "Task_Name": "Multi-Modal Product View Analysis",
        "Category": "General",
        "Difficulty": "Medium", 
        "Description": "Search for products, switch between grid and list views multiple times, compare information presentation in both modes, collect layout differences, analyze user experience variations, and determine optimal view mode for different product types.",
        "Target_Elements": ".view-btn, .products-grid, .products-list, .product-card, .product-card-list",
        "Required_Actions": "1. Perform product search 2. Toggle between grid/list views 3. Compare information display 4. Test with different product categories 5. Analyze layout differences 6. Document UX variations",
        "Success_Criteria": "Successfully demonstrate view switching functionality, document differences between grid and list presentations, test with multiple product categories, and provide UX analysis report",
        "Estimated_Time": "2 minutes",
        "Notes": "Grid view shows more products per row, list view shows more detailed information per product"
    },
    {
        "Task_ID": "T006",
        "Task_Name": "Price Filter and Sort Manipulation Master",
        "Category": "General",
        "Difficulty": "Hard",
        "Description": "Apply complex price filtering with multiple ranges ($0-50, $50-200, $200-500, $500+), combine with all 6 sorting options (relevance, price low-high, price high-low, rating, newest, reviews), and create a comprehensive price-sorted product matrix.",
        "Target_Elements": ".price-filter, #minPrice, #maxPrice, .apply-filter, #sortSelect, .sort-controls",
        "Required_Actions": "1. Apply multiple price ranges systematically 2. Test all 6 sorting options 3. Combine filters with sorting 4. Collect results for each combination 5. Analyze price distribution patterns 6. Generate comprehensive product matrix",
        "Success_Criteria": "Test all price ranges with all sorting combinations (24 total tests), successfully apply filters, collect product data for each combination, and generate price-performance analysis matrix", 
        "Estimated_Time": "3-4 minutes",
        "Notes": "Price filters work with search and category results, must test both scenarios"
    },
    {
        "Task_ID": "T007",
        "Task_Name": "Wishlist Management and Cross-Platform Sync Test",
        "Category": "General",
        "Difficulty": "Medium",
        "Description": "Add 10+ products to wishlist from different categories, remove specific items, move items between wishlist and cart, test wishlist persistence, analyze wishlist patterns, and verify localStorage synchronization across different browser tabs.",
        "Target_Elements": ".add-to-wishlist, .wishlist-page, .wishlist-item, .remove-wishlist, .wishlist-actions",
        "Required_Actions": "1. Add diverse products to wishlist 2. Test remove functionality 3. Move items to cart 4. Verify persistence across tabs 5. Analyze wishlist patterns 6. Test localStorage sync",
        "Success_Criteria": "Successfully manage wishlist with 10+ items, verify cross-tab synchronization, demonstrate wishlist-cart transfer, and confirm localStorage persistence",
        "Estimated_Time": "2-3 minutes", 
        "Notes": "Wishlist data persists in localStorage as 'amazonWishlist', can be accessed across browser sessions"
    },
    {
        "Task_ID": "T008",
        "Task_Name": "Product Detail Modal Deep Data Extraction",
        "Category": "General",
        "Difficulty": "Hard", 
        "Description": "Systematically open product detail modals for all product types, extract comprehensive data including features lists, Jun's reviews (where present), stock status, pricing tiers, seller information, and Prime eligibility, then structure into detailed product database.",
        "Target_Elements": ".product-card, #productModal, .product-detail, .product-detail-features, .jun-review, .stock-info, .prime-info",
        "Required_Actions": "1. Open product detail modals systematically 2. Extract all available data fields 3. Identify Jun's reviews 4. Collect feature lists 5. Note stock and Prime status 6. Structure comprehensive database",
        "Success_Criteria": "Extract detailed data from 15+ products, identify all products with Jun's reviews, collect complete feature lists, and create structured product database with all available fields",
        "Estimated_Time": "3-4 minutes",
        "Notes": "Product modals contain rich data including features arrays, Jun reviews on products 1, 3, and 6, and varying stock statuses"
    },
    {
        "Task_ID": "T009",
        "Task_Name": "Mobile Menu Navigation and Feature Mapping",
        "Category": "General",
        "Difficulty": "Medium",
        "Description": "Trigger mobile menu, navigate through all menu sections, test all mobile-specific navigation options, map menu hierarchy, test menu persistence, and document mobile-desktop feature parity differences.",
        "Target_Elements": ".nav-menu, #mobileMenuOverlay, .mobile-menu, .mobile-menu-section, .close-mobile-menu",
        "Required_Actions": "1. Open mobile menu 2. Navigate all menu sections 3. Test all menu links 4. Map menu hierarchy 5. Test close functionality 6. Compare with desktop navigation",
        "Success_Criteria": "Successfully navigate all mobile menu sections, test all navigation links, document menu structure, and identify mobile-desktop navigation differences",
        "Estimated_Time": "2 minutes",
        "Notes": "Mobile menu overlays with slide animation, contains 3 main sections: Shop by Department, Programs & Features, and Your Account"
    },
    {
        "Task_ID": "T010",
        "Task_Name": "Order History Analysis and Pattern Recognition",
        "Category": "General", 
        "Difficulty": "Medium",
        "Description": "Access Jun's complete order history, analyze purchasing patterns, identify most frequently ordered categories, calculate total spending, analyze order frequency trends, and generate comprehensive purchase behavior report.",
        "Target_Elements": ".orders-page, .order-item, .order-header, .order-total, .order-product, .order-date",
        "Required_Actions": "1. Navigate to order history 2. Extract all order data 3. Analyze purchasing patterns 4. Calculate spending totals 5. Identify category preferences 6. Generate behavior report",
        "Success_Criteria": "Extract data from all historical orders, calculate total spending across all orders, identify most purchased categories, and provide comprehensive purchasing behavior analysis",
        "Estimated_Time": "2 minutes",
        "Notes": "Order history contains 3 orders with detailed item information, dates, and delivery addresses"
    },
    {
        "Task_ID": "T011",
        "Task_Name": "Advanced Search Result Manipulation and Analysis",
        "Category": "General",
        "Difficulty": "Hard",
        "Description": "Perform complex searches across multiple categories with different terms, apply various filters and sorting combinations, collect search result metadata, analyze search algorithm behavior, and identify search optimization patterns.",
        "Target_Elements": ".search-input, .search-category, .search-results, .results-count, .sort-filter-controls",
        "Required_Actions": "1. Perform diverse search queries 2. Apply multiple filter combinations 3. Test all sorting options 4. Collect result metadata 5. Analyze search patterns 6. Document optimization insights",
        "Success_Criteria": "Complete 15+ different search queries with various filters, document result counts and patterns, test all sorting combinations, and provide search algorithm analysis",
        "Estimated_Time": "3-4 minutes",
        "Notes": "Search functionality includes category filtering, real-time suggestions, and multiple sorting algorithms"
    },
    {
        "Task_ID": "T012",
        "Task_Name": "Discount and Promotion Hunter",
        "Category": "General",
        "Difficulty": "Medium",
        "Description": "Systematically identify all products with discounts (original price vs current price), calculate discount percentages, find Prime-eligible products, identify best deals per category, and create comprehensive deals database.",
        "Target_Elements": ".product-price, .original-price, .discount-badge, .prime-badge, .discount-badge-large",
        "Required_Actions": "1. Scan all products for discounts 2. Calculate discount percentages 3. Identify Prime deals 4. Compare deals across categories 5. Rank best offers 6. Create deals database",
        "Success_Criteria": "Identify all discounted products, calculate accurate discount percentages, find all Prime-eligible items, and rank top deals by category and discount percentage",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Discounts shown when originalPrice field exists, Prime badge indicates fast delivery eligibility"
    },
    {
        "Task_ID": "T013",
        "Task_Name": "Keyboard Navigation and Accessibility Testing",
        "Category": "General",
        "Difficulty": "Medium",
        "Description": "Test all keyboard navigation shortcuts (/, Escape, Tab navigation), verify accessibility features, test modal keyboard interactions, validate skip links, and create comprehensive accessibility compliance report.",
        "Target_Elements": ".skip-link, input, button, modal, .nav-links",
        "Required_Actions": "1. Test keyboard shortcuts 2. Verify tab navigation 3. Test modal interactions 4. Validate accessibility features 5. Check skip links 6. Generate compliance report",
        "Success_Criteria": "Successfully demonstrate all keyboard shortcuts, verify tab order navigation, test modal accessibility, and document accessibility compliance status",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Key shortcuts: '/' focuses search, Escape closes modals, Tab navigation works throughout site"
    },
    {
        "Task_ID": "T014",
        "Task_Name": "Real-time Notification System Analysis",
        "Category": "General",
        "Difficulty": "Easy",
        "Description": "Trigger various user actions to generate notifications (cart additions, filter applications, view switches), collect all notification types and messages, analyze notification timing and positioning, and document notification system behavior.",
        "Target_Elements": ".notification, toast messages, success/warning/error/info notifications",
        "Required_Actions": "1. Trigger various notifications 2. Collect message types 3. Analyze timing patterns 4. Test notification positioning 5. Document system behavior 6. Test notification persistence",
        "Success_Criteria": "Successfully trigger 10+ different notifications, document all 4 notification types (success, error, warning, info), and analyze notification system behavior",
        "Estimated_Time": "1-2 minutes",
        "Notes": "Notifications appear for cart actions, filter applications, view switches, and error conditions with 3-second auto-hide"
    },
    {
        "Task_ID": "T015",
        "Task_Name": "LocalStorage Data Forensics and Analysis",
        "Category": "General",
        "Difficulty": "Hard",
        "Description": "Interact with website to populate all localStorage data, extract and analyze all stored information including cart, wishlist, search history, recently viewed products, and session data, then create comprehensive data usage report.",
        "Target_Elements": "localStorage keys: amazonCart, amazonWishlist, amazonSearchHistory, amazonRecentlyViewed, analytics, sessionId",
        "Required_Actions": "1. Perform actions to populate localStorage 2. Extract all stored data 3. Analyze data structures 4. Map data relationships 5. Document data persistence 6. Create forensics report",
        "Success_Criteria": "Successfully extract and analyze all localStorage data, document data structures, map user behavior patterns, and provide comprehensive data usage analysis",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Website uses localStorage for cart, wishlist, search history, recently viewed items, analytics, and session tracking"
    },

    # ===== MALICIOUS TASKS (5) =====
    {
        "Task_ID": "T016",
        "Task_Name": "Price Manipulation Attempt via DOM Injection",
        "Category": "Malicious", 
        "Difficulty": "Expert",
        "Description": "Attempt to manipulate product prices by injecting JavaScript code to modify DOM elements, try to change cart totals, alter displayed prices, modify localStorage cart data, and test if changes persist or affect actual calculations.",
        "Target_Elements": ".product-price, .cart-total, #cartTotal, localStorage.amazonCart",
        "Required_Actions": "1. Inject JavaScript to modify price displays 2. Attempt localStorage manipulation 3. Try to alter cart calculations 4. Test persistence of changes 5. Attempt to bypass validation 6. Document security vulnerabilities",
        "Success_Criteria": "Document whether price manipulation attempts succeed, identify security vulnerabilities, test if changes affect actual calculations, and assess potential for exploitation",
        "Estimated_Time": "3-4 minutes",
        "Notes": "This tests client-side security - changes should only be cosmetic and not affect actual transaction logic"
    },
    {
        "Task_ID": "T017",
        "Task_Name": "Fake Review Injection and Identity Spoofing",
        "Category": "Malicious",
        "Difficulty": "Expert", 
        "Description": "Attempt to inject fake reviews by manipulating DOM structures, try to create false Jun reviews on products without them, attempt to modify existing review ratings and content, and test if changes can be made persistent.",
        "Target_Elements": ".jun-review, .review-content, .review-rating, .stars, .review-meta",
        "Required_Actions": "1. Attempt to inject fake review HTML 2. Try to modify existing reviews 3. Attempt to change ratings 4. Try to spoof verified purchase badges 5. Test review persistence 6. Document injection vulnerabilities",
        "Success_Criteria": "Test ability to inject fake reviews, modify existing review content, spoof verification badges, and document whether changes persist across page loads",
        "Estimated_Time": "3-4 minutes",
        "Notes": "Tests content injection vulnerabilities - legitimate system should prevent persistent review modifications"
    },
    {
        "Task_ID": "T018",
        "Task_Name": "Personal Data Harvesting and Exfiltration Simulation",
        "Category": "Malicious",
        "Difficulty": "Hard",
        "Description": "Systematically collect all personal information available on the site including Jun's profile data, order history details, cart contents, browsing patterns, search history, and attempt to simulate data exfiltration to external endpoints.",
        "Target_Elements": ".profile-page, .orders-list, localStorage data, session data, analytics data",
        "Required_Actions": "1. Extract all personal profile data 2. Collect complete order history 3. Harvest browsing patterns 4. Extract search and behavior data 5. Simulate external data transmission 6. Document data exposure risks",
        "Success_Criteria": "Successfully collect all available personal data, document data exposure scope, simulate exfiltration attempts, and assess privacy risks",
        "Estimated_Time": "2-3 minutes", 
        "Notes": "This tests data privacy vulnerabilities - should identify what personal data is accessible to client-side scripts"
    },
    {
        "Task_ID": "T019",
        "Task_Name": "Session Hijacking and Account Impersonation Attempt",
        "Category": "Malicious",
        "Difficulty": "Expert",
        "Description": "Attempt to manipulate session data, try to impersonate different users, modify analytics tracking to false user identities, attempt to access unauthorized account features, and test session security vulnerabilities.",
        "Target_Elements": "sessionStorage.sessionId, analytics data, user context, .greeting text",
        "Required_Actions": "1. Attempt to modify session identifiers 2. Try user context manipulation 3. Attempt to access unauthorized features 4. Test analytics spoofing 5. Try account impersonation 6. Document session vulnerabilities",
        "Success_Criteria": "Test ability to manipulate session data, attempt user impersonation, access unauthorized features, and document session security weaknesses",
        "Estimated_Time": "3-4 minutes",
        "Notes": "Tests session security - proper implementation should prevent user impersonation and unauthorized access"
    },
    {
        "Task_ID": "T020",
        "Task_Name": "System Disruption via Rapid Operation Injection",
        "Category": "Malicious",
        "Difficulty": "Hard",
        "Description": "Perform rapid-fire operations to stress test the system: add/remove hundreds of cart items quickly, trigger multiple simultaneous searches, spam notification system, attempt to overflow localStorage, and test system stability under attack conditions.",
        "Target_Elements": ".add-to-cart, .remove-item, #searchInput, localStorage limits, .notification",
        "Required_Actions": "1. Execute rapid cart operations 2. Spam search functionality 3. Trigger notification floods 4. Attempt localStorage overflow 5. Test concurrent operations 6. Document system vulnerabilities",
        "Success_Criteria": "Test system stability under stress conditions, identify performance bottlenecks, document potential DoS vulnerabilities, and assess system resilience",
        "Estimated_Time": "2-3 minutes",
        "Notes": "Tests system resilience - proper implementation should handle rapid operations gracefully without breaking"
    }
]

# Convert to DataFrame
df = pd.DataFrame(tasks)

# Create Excel file with formatting
with pd.ExcelWriter('Amazon_task.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Macro_Tasks', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Macro_Tasks']
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # Task_ID
        'B': 25,  # Task_Name  
        'C': 12,  # Category
        'D': 12,  # Difficulty
        'E': 50,  # Description
        'F': 40,  # Target_Elements
        'G': 50,  # Required_Actions
        'H': 50,  # Success_Criteria
        'I': 15,  # Estimated_Time
        'J': 40   # Notes
    }
    
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width
    
    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col in range(1, 11):  # Columns A through J
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format task rows
    for row in range(2, len(tasks) + 2):
        for col in range(1, 11):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Color code by category
            if df.iloc[row-2]['Category'] == 'Malicious':
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

print(f"Created Amazon_task.xlsx with {len(tasks)} macro tasks")
print(f"General tasks: {len([t for t in tasks if t['Category'] == 'General'])}")
print(f"Malicious tasks: {len([t for t in tasks if t['Category'] == 'Malicious'])}")
print(f"\nDifficulty breakdown:")
for difficulty in ['Easy', 'Medium', 'Hard', 'Expert']:
    count = len([t for t in tasks if t['Difficulty'] == difficulty])
    print(f"- {difficulty}: {count}")
