#!/usr/bin/env python3
"""
Excel 파일 생성 스크립트
Airbnb 매크로 task들을 Excel 형태로 변환
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

def create_airbnb_tasks_excel():
    """CSV 파일을 읽어서 고품질 Excel 파일을 생성"""
    
    # CSV 파일 읽기
    df = pd.read_csv('airbnb_tasks.csv')
    
    # Excel 파일 생성
    with pd.ExcelWriter('Airbnb_Tasks.xlsx', engine='openpyxl') as writer:
        # 전체 task 시트
        df.to_excel(writer, sheet_name='All_Tasks', index=False)
        
        # 일반 task 시트 (Malicious_Intent = No)
        general_tasks = df[df['Malicious_Intent'] == 'No']
        general_tasks.to_excel(writer, sheet_name='General_Tasks', index=False)
        
        # 악의적 task 시트 (Malicious_Intent = Yes)
        malicious_tasks = df[df['Malicious_Intent'] == 'Yes']
        malicious_tasks.to_excel(writer, sheet_name='Malicious_Tasks', index=False)
        
        # 난이도별 통계 시트
        difficulty_stats = df.groupby(['Difficulty', 'Malicious_Intent']).size().unstack(fill_value=0)
        difficulty_stats.to_excel(writer, sheet_name='Difficulty_Stats')
    
    # 스타일링 적용
    workbook = openpyxl.load_workbook('Airbnb_Tasks.xlsx')
    
    # 각 시트에 스타일 적용
    for sheet_name in ['All_Tasks', 'General_Tasks', 'Malicious_Tasks']:
        worksheet = workbook[sheet_name]
        style_worksheet(worksheet, sheet_name)
    
    # 통계 시트 스타일링
    stats_sheet = workbook['Difficulty_Stats']
    style_stats_sheet(stats_sheet)
    
    workbook.save('Airbnb_Tasks.xlsx')
    print("✅ Airbnb_Tasks.xlsx 파일이 성공적으로 생성되었습니다!")

def style_worksheet(worksheet, sheet_name):
    """워크시트에 전문적인 스타일 적용"""
    
    # 헤더 스타일
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    
    # 악의적 task 시트는 빨간색 헤더
    if sheet_name == 'Malicious_Tasks':
        header_fill = PatternFill(start_color='C5504B', end_color='C5504B', fill_type='solid')
    
    # 일반 셀 스타일
    cell_font = Font(name='Calibri', size=10)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # 헤더 행 스타일링
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 행 스타일링
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                  min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # 악의적인 task는 연한 빨간색 배경
            if worksheet.cell(row=cell.row, column=9).value == 'Yes':
                cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    
    # 컬럼 너비 자동 조정
    column_widths = {
        'A': 10,  # Task_ID
        'B': 12,  # Category
        'C': 30,  # Task_Title
        'D': 50,  # Task_Description
        'E': 12,  # Difficulty
        'F': 25,  # Target_Elements
        'G': 25,  # Technical_Requirements
        'H': 35,  # Expected_Output
        'I': 15   # Malicious_Intent
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # 행 높이 조정
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 60

def style_stats_sheet(worksheet):
    """통계 시트 스타일링"""
    
    # 헤더 스타일
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # 전체 범위에 테두리 적용
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 헤더 스타일 적용
            if cell.row == 1 or cell.column == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # 컬럼 너비 조정
    for col in ['A', 'B', 'C']:
        worksheet.column_dimensions[col].width = 15

def create_task_summary():
    """Task 요약 정보를 생성"""
    
    summary = {
        'Total Tasks': 20,
        'General Tasks': 15,
        'Malicious Tasks': 5,
        'Difficulty Breakdown': {
            'Easy': 0,
            'Medium': 8,
            'Hard': 12
        },
        'Key Features Tested': [
            'Search & Filtering System',
            'Booking & Reservation Flow',
            'User Profile Management',
            'Host Dashboard Analytics',
            'Review & Rating System',
            'Image Gallery Navigation',
            'Price Analysis & Optimization',
            'Session & Security Management'
        ],
        'Technology Stack': [
            'HTML5/CSS3/JavaScript',
            'Selenium WebDriver',
            'Local Storage API',
            'DOM Manipulation',
            'Data Analysis & Reporting'
        ]
    }
    
    return summary

if __name__ == "__main__":
    print("🚀 Airbnb 매크로 task Excel 파일 생성 중...")
    
    # 현재 디렉토리 확인
    if not os.path.exists('airbnb_tasks.csv'):
        print("❌ airbnb_tasks.csv 파일을 찾을 수 없습니다.")
        print("   스크립트를 Airbnb 폴더에서 실행해주세요.")
        exit(1)
    
    try:
        create_airbnb_tasks_excel()
        
        # 요약 정보 출력
        summary = create_task_summary()
        print(f"\n📊 생성된 Task 요약:")
        print(f"   • 총 Task 수: {summary['Total Tasks']}개")
        print(f"   • 일반 Task: {summary['General Tasks']}개")
        print(f"   • 악의적 Task: {summary['Malicious Tasks']}개")
        print(f"   • 난이도 분포: Medium({summary['Difficulty Breakdown']['Medium']}), Hard({summary['Difficulty Breakdown']['Hard']})")
        
        print(f"\n🎯 주요 테스트 영역:")
        for feature in summary['Key Features Tested']:
            print(f"   • {feature}")
            
        print(f"\n💻 기술 스택:")
        for tech in summary['Technology Stack']:
            print(f"   • {tech}")
            
        print(f"\n📁 생성된 파일: Airbnb_Tasks.xlsx")
        print(f"   • All_Tasks 시트: 전체 20개 task")
        print(f"   • General_Tasks 시트: 일반적인 15개 task") 
        print(f"   • Malicious_Tasks 시트: 악의적인 5개 task")
        print(f"   • Difficulty_Stats 시트: 난이도별 통계")
        
    except Exception as e:
        print(f"❌ Excel 파일 생성 중 오류 발생: {str(e)}")
        print("   필요한 패키지를 설치해주세요: pip install pandas openpyxl")
